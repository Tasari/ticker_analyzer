from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from itertools import islice
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 75 * 1024 * 1024
MAX_ARCHIVE_MEMBERS = 250
DEFAULT_PREVIEW_ROWS = 500
REQUIRED_SHEETS = {"Account Summary"}
ANALYSIS_SHEETS = {"Account Summary", "Account Activity", "Holdings"}
EXTERNAL_FLOW_SUMMARY_KEYS = (
    "Deposits",
    "Internal Transfer",
    "Transfer to Trading (from eToro Money)",
    "Transfer in",
    "Transfer out",
    "Withdrawals",
    "Transfer to eToro Money (from Trading)",
)
FEE_SUMMARY_KEYS = (
    "Overnight Fees",
    "Opening and Closing Spread",
    "Admin Fees",
    "SDRT Charge",
    "Withdrawal Fees",
    "Deposit/Withdrawal FX Conversion Fee",
)


class AccountStatementError(ValueError):
    """Raised when an uploaded workbook is not a supported account statement."""


@dataclass(frozen=True)
class SheetInfo:
    name: str
    data_rows: int
    columns: int


@dataclass(frozen=True)
class StatementOverview:
    currency: str | None
    start_date: datetime | None
    end_date: datetime | None
    sheets: tuple[SheetInfo, ...]


@dataclass(frozen=True)
class SheetPreview:
    columns: tuple[str, ...]
    rows: tuple[tuple[Any, ...], ...]
    total_rows: int

    @property
    def truncated(self) -> bool:
        return self.total_rows > len(self.rows)


@dataclass(frozen=True)
class ExternalCashFlow:
    occurred_at: datetime
    amount: float
    kind: str
    estimated_date: bool = False


@dataclass(frozen=True)
class ExposureGroup:
    name: str
    value: float


@dataclass(frozen=True)
class PositionContribution:
    asset: str
    realized_profit_loss: float
    fees_and_dividends: float
    total_contribution: float
    closed_positions: int


@dataclass(frozen=True)
class StatementAnalysis:
    currency: str
    start_date: datetime
    end_date: datetime
    beginning_realized_equity: float
    ending_realized_equity: float
    beginning_unrealized_equity: float
    ending_unrealized_equity: float
    net_external_flows: float
    positive_contributions: float
    total_profit_loss: float
    closed_positions_profit_loss: float
    dividends: float
    fees: float
    other_performance: float
    unrealized_profit_loss_change: float
    simple_roi: float | None
    annualized_roi: float | None
    modified_dietz_return: float | None
    cash_flows: tuple[ExternalCashFlow, ...]
    open_positions: int
    long_exposure: float
    short_exposure: float
    exposure_by_type: tuple[ExposureGroup, ...]
    warnings: tuple[str, ...]
    holdings_snapshot_date: date | None = None


@dataclass(frozen=True)
class DailyPerformancePoint:
    day: date
    cumulative_profit_loss: float
    estimated_cumulative_profit_loss: float | None = None


@dataclass(frozen=True)
class StatementRangeAnalysis:
    start_date: date
    end_date: date
    realized_profit_loss: float
    closed_positions_profit_loss: float
    dividends: float
    fees: float
    other_performance: float
    net_external_flows: float
    positive_contributions: float
    estimated_beginning_equity: float
    estimated_ending_equity: float
    estimated_total_profit_loss: float
    estimated_roi: float | None
    estimated_annualized_roi: float | None
    estimated_modified_dietz_return: float | None
    holdings_snapshot_count: int
    max_boundary_anchor_distance_days: int
    valuation_warnings: tuple[str, ...]
    daily_performance: tuple[DailyPerformancePoint, ...]


@dataclass(frozen=True)
class _UnrealizedEquityAnchor:
    day: date
    unrealized_profit_loss: float
    exact: bool


def list_statement_assets(payload: bytes) -> tuple[str, ...]:
    """Return stable instrument labels that can be excluded from statement analysis."""
    validate_xlsx_payload(payload)
    try:
        workbook = load_workbook(
            BytesIO(payload),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise AccountStatementError("The uploaded file is not a readable XLSX workbook.") from exc
    try:
        _, assets = _position_asset_index(workbook)
        return tuple(sorted(assets))
    finally:
        workbook.close()


def inspect_account_statement(payload: bytes) -> StatementOverview:
    validate_xlsx_payload(payload)
    try:
        workbook = load_workbook(
            BytesIO(payload),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise AccountStatementError("The uploaded file is not a readable XLSX workbook.") from exc

    try:
        missing = REQUIRED_SHEETS.difference(workbook.sheetnames)
        if missing:
            raise AccountStatementError(
                "This does not look like an eToro account statement: "
                f"missing sheet {', '.join(sorted(missing))}."
            )
        summary = _key_value_rows(workbook["Account Summary"])
        sheets = tuple(
            SheetInfo(name=worksheet.title, data_rows=shape[0], columns=shape[1])
            for worksheet in workbook.worksheets
            for shape in (_worksheet_shape(worksheet),)
        )
        return StatementOverview(
            currency=_optional_text(summary.get("Currency")),
            start_date=_parse_statement_datetime(summary.get("Start Date")),
            end_date=_parse_statement_datetime(summary.get("End Date")),
            sheets=sheets,
        )
    finally:
        workbook.close()


def read_statement_sheet(
    payload: bytes,
    sheet_name: str,
    *,
    max_rows: int = DEFAULT_PREVIEW_ROWS,
) -> SheetPreview:
    if max_rows < 1:
        raise ValueError("max_rows must be positive")
    validate_xlsx_payload(payload)
    try:
        workbook = load_workbook(
            BytesIO(payload),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise AccountStatementError("The uploaded file is not a readable XLSX workbook.") from exc

    try:
        if sheet_name not in workbook.sheetnames:
            raise AccountStatementError(f"Sheet {sheet_name!r} is not present in the workbook.")
        worksheet = workbook[sheet_name]
        total_rows, column_count = _worksheet_shape(worksheet)
        iterator = worksheet.iter_rows(values_only=True)
        first_row = next(iterator, ())
        padded_header = tuple(first_row) + (None,) * max(column_count - len(first_row), 0)
        columns = _unique_headers(padded_header)
        rows = tuple(
            tuple(row[index] if index < len(row) else None for index in range(len(columns)))
            for row in islice(iterator, max_rows)
        )
        return SheetPreview(
            columns=columns,
            rows=rows,
            total_rows=total_rows,
        )
    finally:
        workbook.close()


def analyze_account_statement(
    payload: bytes,
    *,
    excluded_assets: tuple[str, ...] = (),
) -> StatementAnalysis:
    validate_xlsx_payload(payload)
    try:
        workbook = load_workbook(
            BytesIO(payload),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise AccountStatementError("The uploaded file is not a readable XLSX workbook.") from exc

    try:
        missing_sheets = ANALYSIS_SHEETS.difference(workbook.sheetnames)
        if missing_sheets:
            raise AccountStatementError(
                "Portfolio analysis is unavailable: missing sheet "
                f"{', '.join(sorted(missing_sheets))}."
            )
        summary = _key_value_rows(workbook["Account Summary"])
        start_date = _required_datetime(summary, "Start Date")
        end_date = _required_datetime(summary, "End Date")
        if end_date <= start_date:
            raise AccountStatementError("The statement end date must be after its start date.")

        beginning_realized = _required_number(summary, "Beginning Realized Equity")
        ending_realized = _required_number(summary, "Ending Realized Equity")
        beginning_unrealized = _required_number(summary, "Beginning Unrealized Equity")
        ending_unrealized = _required_number(summary, "Ending Unrealized Equity")
        summary_external_flows = sum(_number(summary.get(key)) for key in EXTERNAL_FLOW_SUMMARY_KEYS)
        cash_flows = list(_external_cash_flows(workbook["Account Activity"]))
        warnings: list[str] = []
        dated_total = sum(flow.amount for flow in cash_flows)
        undated_flow = summary_external_flows - dated_total
        if abs(undated_flow) >= 0.005:
            cash_flows.append(
                ExternalCashFlow(
                    occurred_at=start_date + (end_date - start_date) / 2,
                    amount=undated_flow,
                    kind="Undated statement cash flow",
                    estimated_date=True,
                )
            )
            warnings.append(
                "Some external cash flows had no matching activity date and were weighted at the period midpoint."
            )

        cash_flows.sort(key=lambda flow: flow.occurred_at)
        net_external_flows = sum(flow.amount for flow in cash_flows)
        positive_contributions = sum(max(flow.amount, 0.0) for flow in cash_flows)
        total_profit_loss = ending_unrealized - beginning_unrealized - net_external_flows
        closed_profit_loss = _number(summary.get("Profit or Loss (Closed positions only)"))
        dividends = _number(summary.get("Dividends")) + _number(summary.get("Dividend CFD"))
        fees = sum(_number(summary.get(key)) for key in FEE_SUMMARY_KEYS)
        unrealized_change = (
            ending_unrealized - ending_realized
        ) - (beginning_unrealized - beginning_realized)
        other_performance = total_profit_loss - closed_profit_loss - dividends - fees - unrealized_change

        invested_capital = beginning_unrealized + positive_contributions
        simple_roi = total_profit_loss / invested_capital if invested_capital > 0 else None
        annualized_roi = annualize_return(simple_roi, start_date, end_date)
        dietz = modified_dietz_return(
            beginning_unrealized,
            ending_unrealized,
            cash_flows,
            start_date,
            end_date,
        )
        position_assets, _ = _position_asset_index(workbook)
        excluded = {_normalize_asset(value) for value in excluded_assets}
        (
            open_positions,
            long_exposure,
            short_exposure,
            exposure_by_type,
            holdings_snapshot_date,
        ) = _holdings_exposure(
            workbook["Holdings"],
            position_assets=position_assets,
            excluded_assets=excluded,
        )
        return StatementAnalysis(
            currency=_optional_text(summary.get("Currency")) or "USD",
            start_date=start_date,
            end_date=end_date,
            beginning_realized_equity=beginning_realized,
            ending_realized_equity=ending_realized,
            beginning_unrealized_equity=beginning_unrealized,
            ending_unrealized_equity=ending_unrealized,
            net_external_flows=net_external_flows,
            positive_contributions=positive_contributions,
            total_profit_loss=total_profit_loss,
            closed_positions_profit_loss=closed_profit_loss,
            dividends=dividends,
            fees=fees,
            other_performance=other_performance,
            unrealized_profit_loss_change=unrealized_change,
            simple_roi=simple_roi,
            annualized_roi=annualized_roi,
            modified_dietz_return=dietz,
            cash_flows=tuple(cash_flows),
            open_positions=open_positions,
            long_exposure=long_exposure,
            short_exposure=short_exposure,
            exposure_by_type=exposure_by_type,
            warnings=tuple(warnings),
            holdings_snapshot_date=holdings_snapshot_date,
        )
    finally:
        workbook.close()


def modified_dietz_return(
    beginning_value: float,
    ending_value: float,
    cash_flows: list[ExternalCashFlow] | tuple[ExternalCashFlow, ...],
    start_date: datetime,
    end_date: datetime,
) -> float | None:
    duration = (end_date - start_date).total_seconds()
    if duration <= 0:
        return None
    weighted_flows = 0.0
    total_flows = 0.0
    for flow in cash_flows:
        remaining = (end_date - flow.occurred_at).total_seconds()
        weight = min(max(remaining / duration, 0.0), 1.0)
        weighted_flows += weight * flow.amount
        total_flows += flow.amount
    denominator = beginning_value + weighted_flows
    if denominator <= 0:
        return None
    return (ending_value - beginning_value - total_flows) / denominator


def analyze_statement_range(
    payload: bytes,
    start_date: date,
    end_date: date,
    *,
    excluded_assets: tuple[str, ...] = (),
) -> StatementRangeAnalysis:
    validate_xlsx_payload(payload)
    if end_date < start_date:
        raise AccountStatementError("The selected end date must not be before the start date.")
    try:
        workbook = load_workbook(
            BytesIO(payload),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise AccountStatementError("The uploaded file is not a readable XLSX workbook.") from exc

    try:
        missing_range_sheets = {"Account Activity", "Holdings"}.difference(workbook.sheetnames)
        if missing_range_sheets:
            raise AccountStatementError(
                "Date-range analysis is unavailable: missing sheet "
                f"{', '.join(sorted(missing_range_sheets))}."
            )
        summary = _key_value_rows(workbook["Account Summary"])
        statement_start = _required_datetime(summary, "Start Date").date()
        statement_end = _required_datetime(summary, "End Date").date()
        if start_date < statement_start or end_date > statement_end:
            raise AccountStatementError(
                "The selected dates must stay within the account statement period."
            )
        activity = workbook["Account Activity"]
        position_assets, _ = _position_asset_index(workbook)
        excluded = {_normalize_asset(value) for value in excluded_assets}
        range_analysis = _range_activity_analysis(
            activity,
            start_date,
            end_date,
            position_assets=position_assets,
            excluded_assets=excluded,
        )
        statement_start_at = _required_datetime(summary, "Start Date")
        statement_end_at = _required_datetime(summary, "End Date")
        beginning_realized = _required_number(summary, "Beginning Realized Equity")
        ending_realized = _required_number(summary, "Ending Realized Equity")
        beginning_unrealized = _required_number(summary, "Beginning Unrealized Equity")
        ending_unrealized = _required_number(summary, "Ending Unrealized Equity")
        anchors, snapshot_count, valuation_warnings = _unrealized_equity_anchors(
            workbook["Holdings"],
            statement_start_at.date(),
            statement_end_at.date(),
            beginning_unrealized - beginning_realized,
            ending_unrealized - ending_realized,
        )
        start_unrealized, start_distance = _interpolate_unrealized(start_date, anchors)
        end_unrealized, end_distance = _interpolate_unrealized(end_date, anchors)
        start_realized = _realized_equity_at(
            activity,
            start_date,
            inclusive=False,
            fallback=beginning_realized,
        )
        end_realized = _realized_equity_at(
            activity,
            end_date,
            inclusive=True,
            fallback=beginning_realized,
        )
        if start_date == statement_start_at.date():
            start_realized = beginning_realized
        if end_date == statement_end_at.date():
            end_realized = ending_realized
        estimated_beginning_equity = start_realized + start_unrealized
        estimated_ending_equity = end_realized + end_unrealized
        cash_flows = tuple(
            flow
            for flow in _external_cash_flows(activity)
            if start_date <= flow.occurred_at.date() <= end_date
        )
        positive_contributions = sum(max(flow.amount, 0.0) for flow in cash_flows)
        if excluded:
            estimated_profit_loss = range_analysis.realized_profit_loss + end_unrealized - start_unrealized
            estimated_ending_equity = (
                estimated_beginning_equity
                + range_analysis.net_external_flows
                + estimated_profit_loss
            )
            valuation_warnings = (
                *valuation_warnings,
                "Excluded instruments are removed from Position-ID-linked realized activity. "
                "Historical aggregate unrealized valuations cannot be separated by instrument and remain estimated.",
            )
        else:
            estimated_profit_loss = (
                estimated_ending_equity
                - estimated_beginning_equity
                - range_analysis.net_external_flows
            )
        invested_capital = estimated_beginning_equity + positive_contributions
        estimated_roi = estimated_profit_loss / invested_capital if invested_capital > 0 else None
        period_start_at = datetime.combine(start_date, datetime.min.time())
        period_end_at = datetime.combine(end_date + timedelta(days=1), datetime.min.time())
        estimated_annualized_roi = annualize_return(
            estimated_roi,
            period_start_at,
            period_end_at,
        )
        estimated_dietz = modified_dietz_return(
            estimated_beginning_equity,
            estimated_ending_equity,
            cash_flows,
            period_start_at,
            period_end_at,
        )
        daily_performance = _estimated_daily_performance(
            activity,
            start_date,
            end_date,
            start_realized,
            estimated_beginning_equity,
            anchors,
            range_analysis.daily_performance,
            filtered=bool(excluded),
        )
        return StatementRangeAnalysis(
            start_date=range_analysis.start_date,
            end_date=range_analysis.end_date,
            realized_profit_loss=range_analysis.realized_profit_loss,
            closed_positions_profit_loss=range_analysis.closed_positions_profit_loss,
            dividends=range_analysis.dividends,
            fees=range_analysis.fees,
            other_performance=range_analysis.other_performance,
            net_external_flows=range_analysis.net_external_flows,
            positive_contributions=positive_contributions,
            estimated_beginning_equity=estimated_beginning_equity,
            estimated_ending_equity=estimated_ending_equity,
            estimated_total_profit_loss=estimated_profit_loss,
            estimated_roi=estimated_roi,
            estimated_annualized_roi=estimated_annualized_roi,
            estimated_modified_dietz_return=estimated_dietz,
            holdings_snapshot_count=snapshot_count,
            max_boundary_anchor_distance_days=max(start_distance, end_distance),
            valuation_warnings=valuation_warnings,
            daily_performance=daily_performance,
        )
    finally:
        workbook.close()


def analyze_position_contributions(
    payload: bytes,
    start_date: date,
    end_date: date,
    *,
    excluded_assets: tuple[str, ...] = (),
) -> tuple[PositionContribution, ...]:
    """Aggregate exact closed-position results by asset for the selected close-date range."""
    validate_xlsx_payload(payload)
    if end_date < start_date:
        raise AccountStatementError("The selected end date must not be before the start date.")
    try:
        workbook = load_workbook(
            BytesIO(payload),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise AccountStatementError("The uploaded file is not a readable XLSX workbook.") from exc
    try:
        if "Closed Positions" not in workbook.sheetnames:
            return ()
        rows = workbook["Closed Positions"].iter_rows(values_only=True)
        header = next(rows, ())
        columns = {_optional_text(value): index for index, value in enumerate(header)}
        required = {"Action", "Close Date", "Profit(USD)"}
        if not required.issubset(columns):
            return ()
        excluded = {_normalize_asset(value) for value in excluded_assets}
        totals: dict[str, list[float]] = {}
        extra_index = columns.get("Overnight Fees and Dividends")
        for row in rows:
            closed_at = _parse_statement_datetime(_row_value(row, columns["Close Date"]))
            if closed_at is None or not start_date <= closed_at.date() <= end_date:
                continue
            asset = _optional_text(_row_value(row, columns["Action"])) or "Unknown"
            if _normalize_asset(_asset_label(asset)) in excluded:
                continue
            profit = _number(_row_value(row, columns["Profit(USD)"]))
            extras = _number(_row_value(row, extra_index)) if extra_index is not None else 0.0
            aggregate = totals.setdefault(asset, [0.0, 0.0, 0.0])
            aggregate[0] += profit
            aggregate[1] += extras
            aggregate[2] += 1
        return tuple(
            PositionContribution(
                asset=asset,
                realized_profit_loss=values[0],
                fees_and_dividends=values[1],
                # eToro's Profit(USD) already reconciles to the statement's
                # closed-position P/L. The fee/dividend column is a component,
                # not an additional amount.
                total_contribution=values[0],
                closed_positions=int(values[2]),
            )
            for asset, values in sorted(
                totals.items(),
                key=lambda item: abs(item[1][0]),
                reverse=True,
            )
        )
    finally:
        workbook.close()


def annualize_return(
    period_return: float | None,
    start_date: datetime,
    end_date: datetime,
) -> float | None:
    if period_return is None or period_return <= -1:
        return None
    years = (end_date - start_date).total_seconds() / timedelta(days=365.2425).total_seconds()
    if years <= 0:
        return None
    return (1 + period_return) ** (1 / years) - 1


def validate_xlsx_payload(payload: bytes) -> None:
    if not payload:
        raise AccountStatementError("The uploaded file is empty.")
    if len(payload) > MAX_UPLOAD_BYTES:
        raise AccountStatementError("The workbook is larger than the 10 MB upload limit.")
    try:
        with ZipFile(BytesIO(payload)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ARCHIVE_MEMBERS:
                raise AccountStatementError("The workbook contains too many internal files.")
            if sum(member.file_size for member in members) > MAX_UNCOMPRESSED_BYTES:
                raise AccountStatementError("The expanded workbook is too large to process safely.")
            if "xl/workbook.xml" not in archive.namelist():
                raise AccountStatementError("The uploaded file is not an XLSX workbook.")
    except BadZipFile as exc:
        raise AccountStatementError("The uploaded file is not an XLSX workbook.") from exc


def _key_value_rows(worksheet: Any) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for row in worksheet.iter_rows(min_col=1, max_col=2, values_only=True):
        key = _optional_text(row[0])
        if key:
            values[key] = row[1]
    return values


def _external_cash_flows(worksheet: Any) -> tuple[ExternalCashFlow, ...]:
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, ())
    columns = {_optional_text(value): index for index, value in enumerate(header)}
    required = {"Date", "Type", "Amount"}
    if not required.issubset(columns):
        return ()
    flows: list[ExternalCashFlow] = []
    equity_change_index = columns.get("Realized Equity Change")
    for row in rows:
        kind = _optional_text(_row_value(row, columns["Type"]))
        if not kind or not _is_external_flow(kind):
            continue
        occurred_at = _parse_statement_datetime(_row_value(row, columns["Date"]))
        if occurred_at is None:
            continue
        amount = _number(_row_value(row, columns["Amount"]))
        if equity_change_index is not None:
            equity_change = _number(_row_value(row, equity_change_index))
            if abs(equity_change) >= 0.005:
                amount = equity_change
        if abs(amount) >= 0.005:
            flows.append(ExternalCashFlow(occurred_at=occurred_at, amount=amount, kind=kind))
    return tuple(flows)


def _range_activity_analysis(
    worksheet: Any,
    start_date: date,
    end_date: date,
    *,
    position_assets: dict[str, str] | None = None,
    excluded_assets: set[str] | None = None,
) -> StatementRangeAnalysis:
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, ())
    columns = {_optional_text(value): index for index, value in enumerate(header)}
    required = {"Date", "Type", "Amount", "Realized Equity Change"}
    if not required.issubset(columns):
        raise AccountStatementError(
            "Date-range analysis is unavailable: Account Activity has unsupported columns."
        )

    closed_profit_loss = 0.0
    dividends = 0.0
    fees = 0.0
    other_performance = 0.0
    net_external_flows = 0.0
    realized_profit_loss = 0.0
    daily_changes: dict[date, float] = {}
    position_index = columns.get("Position ID")
    for row in rows:
        occurred_at = _parse_statement_datetime(_row_value(row, columns["Date"]))
        if occurred_at is None or not start_date <= occurred_at.date() <= end_date:
            continue
        kind = _optional_text(_row_value(row, columns["Type"])) or "Unknown"
        if position_index is not None and excluded_assets:
            position_id = _position_id(_row_value(row, position_index))
            asset = (position_assets or {}).get(position_id)
            if asset and _normalize_asset(asset) in excluded_assets:
                continue
        amount = _number(_row_value(row, columns["Amount"]))
        equity_change = _number(_row_value(row, columns["Realized Equity Change"]))
        normalized = kind.casefold()
        performance_change = 0.0
        if _is_external_flow(kind):
            net_external_flows += equity_change if abs(equity_change) >= 0.005 else amount
        elif normalized == "position closed":
            closed_profit_loss += equity_change
            performance_change = equity_change
        elif normalized == "dividend":
            dividends += equity_change
            performance_change = equity_change
        elif _is_fee_activity(kind):
            fee_change = amount if abs(amount) >= 0.005 else equity_change
            fees += fee_change
            performance_change = equity_change
        elif not _is_non_performance_activity(kind):
            other_performance += equity_change
            performance_change = equity_change
        if not _is_external_flow(kind) and not _is_non_performance_activity(kind):
            realized_profit_loss += equity_change
        if abs(performance_change) >= 0.005:
            day = occurred_at.date()
            daily_changes[day] = daily_changes.get(day, 0.0) + performance_change

    cumulative = 0.0
    daily_performance: list[DailyPerformancePoint] = []
    current_day = start_date
    while current_day <= end_date:
        cumulative += daily_changes.get(current_day, 0.0)
        daily_performance.append(
            DailyPerformancePoint(day=current_day, cumulative_profit_loss=cumulative)
        )
        current_day += timedelta(days=1)
    other_performance = realized_profit_loss - closed_profit_loss - dividends - fees
    return StatementRangeAnalysis(
        start_date=start_date,
        end_date=end_date,
        realized_profit_loss=realized_profit_loss,
        closed_positions_profit_loss=closed_profit_loss,
        dividends=dividends,
        fees=fees,
        other_performance=other_performance,
        net_external_flows=net_external_flows,
        positive_contributions=0.0,
        estimated_beginning_equity=0.0,
        estimated_ending_equity=0.0,
        estimated_total_profit_loss=0.0,
        estimated_roi=None,
        estimated_annualized_roi=None,
        estimated_modified_dietz_return=None,
        holdings_snapshot_count=0,
        max_boundary_anchor_distance_days=0,
        valuation_warnings=(),
        daily_performance=tuple(daily_performance),
    )


def _unrealized_equity_anchors(
    worksheet: Any,
    statement_start: date,
    statement_end: date,
    beginning_unrealized_profit_loss: float,
    ending_unrealized_profit_loss: float,
) -> tuple[tuple[_UnrealizedEquityAnchor, ...], int, tuple[str, ...]]:
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, ())
    columns = {_optional_text(value): index for index, value in enumerate(header)}
    required = {"Snapshot Date", "Direction", "Open Rate", "Current Rate", "Value in USD"}
    snapshot_totals: dict[date, float] = {}
    snapshot_exposure: dict[date, float] = {}
    covered_exposure: dict[date, float] = {}
    if required.issubset(columns):
        for row in rows:
            snapshot_at = _parse_statement_datetime(_row_value(row, columns["Snapshot Date"]))
            if snapshot_at is None:
                continue
            day = snapshot_at.date()
            value = abs(_number(_row_value(row, columns["Value in USD"])))
            snapshot_exposure[day] = snapshot_exposure.get(day, 0.0) + value
            open_rate = _number(_row_value(row, columns["Open Rate"]))
            current_rate = _number(_row_value(row, columns["Current Rate"]))
            if value <= 0 or open_rate <= 0 or current_rate <= 0:
                continue
            direction = (
                _optional_text(_row_value(row, columns["Direction"])) or "Long"
            ).casefold()
            if direction == "short":
                profit_loss = value * (open_rate / current_rate - 1)
            else:
                profit_loss = value * (1 - open_rate / current_rate)
            snapshot_totals[day] = snapshot_totals.get(day, 0.0) + profit_loss
            covered_exposure[day] = covered_exposure.get(day, 0.0) + value

    anchors_by_day = {
        day: _UnrealizedEquityAnchor(day, value, exact=False)
        for day, value in snapshot_totals.items()
    }
    anchors_by_day[statement_start] = _UnrealizedEquityAnchor(
        statement_start,
        beginning_unrealized_profit_loss,
        exact=True,
    )
    anchors_by_day[statement_end] = _UnrealizedEquityAnchor(
        statement_end,
        ending_unrealized_profit_loss,
        exact=True,
    )
    warnings: list[str] = []
    for day, exposure in snapshot_exposure.items():
        coverage = covered_exposure.get(day, 0.0) / exposure if exposure > 0 else 1.0
        if coverage < 0.9:
            warnings.append(
                f"Holdings valuation coverage on {day:%Y-%m-%d} was {coverage:.0%}."
            )
    intermediate_snapshot_count = sum(
        day not in {statement_start, statement_end} for day in snapshot_totals
    )
    if not intermediate_snapshot_count:
        warnings.append(
            "No usable intermediate Holdings snapshots were found; valuation is interpolated "
            "between the statement boundaries."
        )
    return (
        tuple(sorted(anchors_by_day.values(), key=lambda anchor: anchor.day)),
        intermediate_snapshot_count,
        tuple(warnings),
    )


def _interpolate_unrealized(
    target: date,
    anchors: tuple[_UnrealizedEquityAnchor, ...],
) -> tuple[float, int]:
    before = max((anchor for anchor in anchors if anchor.day <= target), key=lambda item: item.day)
    after = min((anchor for anchor in anchors if anchor.day >= target), key=lambda item: item.day)
    nearest_distance = min(abs((target - before.day).days), abs((after.day - target).days))
    duration = (after.day - before.day).days
    if duration <= 0:
        return before.unrealized_profit_loss, nearest_distance
    elapsed = (target - before.day).days
    weight = elapsed / duration
    value = before.unrealized_profit_loss + weight * (
        after.unrealized_profit_loss - before.unrealized_profit_loss
    )
    return value, nearest_distance


def _realized_equity_at(
    worksheet: Any,
    target: date,
    *,
    inclusive: bool,
    fallback: float,
) -> float:
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, ())
    columns = {_optional_text(value): index for index, value in enumerate(header)}
    if not {"Date", "Realized Equity"}.issubset(columns):
        return fallback
    result = fallback
    for row in rows:
        occurred_at = _parse_statement_datetime(_row_value(row, columns["Date"]))
        if occurred_at is None:
            continue
        in_boundary = occurred_at.date() <= target if inclusive else occurred_at.date() < target
        if in_boundary:
            result = _number(_row_value(row, columns["Realized Equity"]))
    return result


def _estimated_daily_performance(
    worksheet: Any,
    start_date: date,
    end_date: date,
    starting_realized_equity: float,
    starting_total_equity: float,
    anchors: tuple[_UnrealizedEquityAnchor, ...],
    realized_points: tuple[DailyPerformancePoint, ...],
    *,
    filtered: bool = False,
) -> tuple[DailyPerformancePoint, ...]:
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, ())
    columns = {_optional_text(value): index for index, value in enumerate(header)}
    realized_by_day: dict[date, float] = {}
    external_flows_by_day: dict[date, float] = {}
    if {"Date", "Type", "Amount", "Realized Equity", "Realized Equity Change"}.issubset(columns):
        for row in rows:
            occurred_at = _parse_statement_datetime(_row_value(row, columns["Date"]))
            if occurred_at is None or not start_date <= occurred_at.date() <= end_date:
                continue
            day = occurred_at.date()
            realized_by_day[day] = _number(_row_value(row, columns["Realized Equity"]))
            kind = _optional_text(_row_value(row, columns["Type"])) or ""
            if _is_external_flow(kind):
                equity_change = _number(_row_value(row, columns["Realized Equity Change"]))
                amount = _number(_row_value(row, columns["Amount"]))
                external_flows_by_day[day] = external_flows_by_day.get(day, 0.0) + (
                    equity_change if abs(equity_change) >= 0.005 else amount
                )
    current_realized = starting_realized_equity
    cumulative_flows = 0.0
    starting_unrealized, _ = _interpolate_unrealized(start_date, anchors)
    points: list[DailyPerformancePoint] = []
    for point in realized_points:
        current_realized = realized_by_day.get(point.day, current_realized)
        cumulative_flows += external_flows_by_day.get(point.day, 0.0)
        unrealized, _ = _interpolate_unrealized(point.day, anchors)
        if filtered:
            estimated_total_profit_loss = point.cumulative_profit_loss + unrealized - starting_unrealized
        else:
            estimated_total_profit_loss = (
                current_realized + unrealized - starting_total_equity - cumulative_flows
            )
        points.append(
            DailyPerformancePoint(
                day=point.day,
                cumulative_profit_loss=point.cumulative_profit_loss,
                estimated_cumulative_profit_loss=estimated_total_profit_loss,
            )
        )
    return tuple(points)


def _is_external_flow(kind: str) -> bool:
    normalized = kind.casefold()
    if "fee" in normalized or "mirror" in normalized or "copy" in normalized:
        return False
    return (
        normalized in {"deposit", "withdrawal", "transfer in", "transfer out", "internal transfer"}
        or "etoro money" in normalized
    )


def _is_fee_activity(kind: str) -> bool:
    normalized = kind.casefold()
    return "fee" in normalized or normalized in {"sdrt", "admin fee"}


def _is_non_performance_activity(kind: str) -> bool:
    normalized = kind.casefold()
    return (
        normalized == "open position"
        or "mirror" in normalized
        or "copy" in normalized
        or normalized.startswith("corp action")
    )


def _holdings_exposure(
    worksheet: Any,
    *,
    position_assets: dict[str, str] | None = None,
    excluded_assets: set[str] | None = None,
) -> tuple[int, float, float, tuple[ExposureGroup, ...], date | None]:
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, ())
    columns = {_optional_text(value): index for index, value in enumerate(header)}
    required = {"Direction", "Value in USD"}
    if not required.issubset(columns):
        return 0, 0.0, 0.0, (), None
    data_rows = list(rows)
    snapshot_date: date | None = None
    snapshot_index = columns.get("Snapshot Date")
    if snapshot_index is not None:
        dated_rows = [
            (parsed.date(), row)
            for row in data_rows
            for parsed in (_parse_statement_datetime(_row_value(row, snapshot_index)),)
            if parsed is not None
        ]
        if dated_rows:
            snapshot_date = max(day for day, _ in dated_rows)
            data_rows = [row for day, row in dated_rows if day == snapshot_date]
    positions = 0
    long_exposure = 0.0
    short_exposure = 0.0
    by_type: dict[str, float] = {}
    type_index = columns.get("Type")
    position_index = columns.get("Position ID")
    for row in data_rows:
        if position_index is not None and excluded_assets:
            position_id = _position_id(_row_value(row, position_index))
            asset = (position_assets or {}).get(position_id)
            if asset and _normalize_asset(asset) in excluded_assets:
                continue
        raw_value = _row_value(row, columns["Value in USD"])
        if raw_value is None:
            continue
        exposure = abs(_number(raw_value))
        direction = (_optional_text(_row_value(row, columns["Direction"])) or "Long").casefold()
        asset_type = (
            _optional_text(_row_value(row, type_index)) if type_index is not None else None
        ) or "Unknown"
        positions += 1
        if direction == "short":
            short_exposure += exposure
        else:
            long_exposure += exposure
        by_type[asset_type] = by_type.get(asset_type, 0.0) + exposure
    groups = tuple(
        ExposureGroup(name=name, value=value)
        for name, value in sorted(by_type.items(), key=lambda item: item[1], reverse=True)
    )
    return positions, long_exposure, short_exposure, groups, snapshot_date


def _position_asset_index(workbook: Any) -> tuple[dict[str, str], set[str]]:
    position_assets: dict[str, str] = {}
    assets: set[str] = set()
    name_aliases: dict[str, str] = {}
    if "Closed Positions" in workbook.sheetnames:
        rows = workbook["Closed Positions"].iter_rows(values_only=True)
        columns = {_optional_text(value): index for index, value in enumerate(next(rows, ()))}
        if "Action" in columns:
            for row in rows:
                action = _optional_text(_row_value(row, columns["Action"]))
                if not action:
                    continue
                label = _asset_label(action)
                assets.add(label)
                name_aliases[_normalize_asset(_asset_name(action))] = label
                if "Position ID" in columns:
                    position_assets[_position_id(_row_value(row, columns["Position ID"]))] = label
    if "Holdings" in workbook.sheetnames:
        rows = workbook["Holdings"].iter_rows(values_only=True)
        columns = {_optional_text(value): index for index, value in enumerate(next(rows, ()))}
        if "Asset" in columns:
            for row in rows:
                asset = _optional_text(_row_value(row, columns["Asset"]))
                if not asset:
                    continue
                label = name_aliases.get(_normalize_asset(asset), _asset_label(asset))
                assets.add(label)
                if "Position ID" in columns:
                    position_assets.setdefault(
                        _position_id(_row_value(row, columns["Position ID"])),
                        label,
                    )
    position_assets.pop("", None)
    return position_assets, assets


_TICKER_SUFFIX = re.compile(r"\(([A-Z0-9.^=\-]{1,20})\)\s*$")


def _asset_label(value: str) -> str:
    match = _TICKER_SUFFIX.search(value.strip())
    return match.group(1) if match else value.strip()


def _asset_name(value: str) -> str:
    return _TICKER_SUFFIX.sub("", value).strip()


def _normalize_asset(value: str) -> str:
    return value.strip().casefold()


def _position_id(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() if value is not None else ""


def _row_value(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def _required_datetime(summary: dict[str, Any], key: str) -> datetime:
    value = _parse_statement_datetime(summary.get(key))
    if value is None:
        raise AccountStatementError(f"Portfolio analysis is unavailable: missing {key}.")
    return value


def _required_number(summary: dict[str, Any], key: str) -> float:
    if summary.get(key) is None:
        raise AccountStatementError(f"Portfolio analysis is unavailable: missing {key}.")
    return _number(summary[key])


def _number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = _optional_text(value)
    if not text or text == "-":
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    normalized = text.strip("()").replace(",", "").replace("$", "")
    try:
        parsed = float(normalized)
    except ValueError:
        return 0.0
    return -parsed if negative else parsed


def _worksheet_shape(worksheet: Any) -> tuple[int, int]:
    if worksheet.max_row and worksheet.max_column:
        return max(worksheet.max_row - 1, 0), worksheet.max_column
    row_count = 0
    column_count = 0
    for row in worksheet.iter_rows(values_only=True):
        row_count += 1
        column_count = max(column_count, len(row))
    return max(row_count - 1, 0), column_count


def _unique_headers(values: tuple[Any, ...]) -> tuple[str, ...]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = _optional_text(value) or f"Column {index}"
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base} ({counts[base]})")
    return tuple(headers)


def _optional_text(value: Any) -> str | None:
    text = str(value).strip() if value is not None else ""
    return text or None


def _parse_statement_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = _optional_text(value)
    if not text:
        return None
    for pattern in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None
