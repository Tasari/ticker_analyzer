from __future__ import annotations

import unittest
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from ticker_analyzer.account_statement import (
    AccountStatementError,
    ExternalCashFlow,
    analyze_account_statement,
    analyze_position_contributions,
    analyze_statement_range,
    annualize_return,
    inspect_account_statement,
    list_statement_assets,
    modified_dietz_return,
    read_statement_sheet,
)


def statement_workbook(*, include_summary: bool = True) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Account Summary" if include_summary else "Other"
    summary.append(["Details", "Totals"])
    summary.append(["Currency", "USD"])
    summary.append(["Start Date", "01/06/2023 00:00:00"])
    summary.append(["End Date", "31/08/2023 23:59:59"])
    holdings = workbook.create_sheet("Holdings")
    holdings.append(["Asset", "Position ID", "Value in USD"])
    holdings.append(["Apple", 123, 10.5])
    holdings.append(["Bitcoin", 456, 20.25])
    glossary = workbook.create_sheet("Glossary")
    glossary.append(["Account Statement Glossary"])
    glossary.append(["Term", "Meaning"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def analysis_workbook(
    *,
    dated_deposit: bool = True,
    holdings_snapshot: bool = False,
    include_gold: bool = False,
) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Account Summary"
    rows = [
        ("Details", "Totals"),
        ("Currency", "USD"),
        ("Start Date", "01/01/2024 00:00:00"),
        ("End Date", "11/01/2024 00:00:00"),
        ("Beginning Realized Equity", 100),
        ("Ending Realized Equity", 165),
        ("Beginning Unrealized Equity", 100),
        ("Ending Unrealized Equity", 165),
        ("Deposits", 50),
        ("Profit or Loss (Closed positions only)", 5),
        ("Dividends", 2),
        ("Dividend CFD", 0),
        ("Overnight Fees", -1),
    ]
    for row in rows:
        summary.append(row)

    activity = workbook.create_sheet("Account Activity")
    activity.append(["Date", "Type", "Amount", "Realized Equity Change", "Realized Equity", "Position ID"])
    if dated_deposit:
        activity.append(["06/01/2024 00:00:00", "Deposit", 50, 50, 150, None])
    activity.append(["07/01/2024 12:00:00", "Position closed", 25, 5, 155, 101])
    activity.append(["08/01/2024 12:00:00", "Dividend", 2, 2, 157, 101])
    activity.append(["09/01/2024 12:00:00", "Overnight fee", -1, -1, 156, 101])
    if include_gold:
        activity.append(["09/01/2024 15:00:00", "Position closed", 10, 10, 166, 202])
    activity.append(["10/01/2024 12:00:00", "Open Position", 20, 0, 156, 101])

    holdings = workbook.create_sheet("Holdings")
    if holdings_snapshot:
        holdings.append(
            [
                "Snapshot Date", "Asset", "Position ID", "Direction", "Value in USD", "Type",
                "Open Rate", "Current Rate",
            ]
        )
        holdings.append([datetime(2024, 1, 6), "Alpha", 101, "Long", 120, "Stocks", 100, 120])
        holdings.append([datetime(2024, 1, 6), "Beta", 102, "Short", -60, "CFD", 120, 100])
    else:
        holdings.append(["Asset", "Position ID", "Direction", "Value in USD", "Type"])
        holdings.append(["Alpha", 101, "Long", 80, "Stocks"])
        holdings.append(["Beta", 102, "Short", -20, "CFD"])
    if include_gold:
        if holdings_snapshot:
            holdings.append(
                [datetime(2024, 1, 6), "Gold (Non Expiry)", 202, "Long", 40, "Commodities", 100, 110]
            )
        else:
            holdings.append(["Gold (Non Expiry)", 202, "Long", 40, "Commodities"])
    closed = workbook.create_sheet("Closed Positions")
    closed.append(["Action", "Close Date", "Profit(USD)", "Overnight Fees and Dividends", "Position ID"])
    closed.append(["Alpha (AAA)", "08/01/2024 12:00:00", 5, -0.5, 101])
    closed.append(["Alpha (AAA)", "10/01/2024 12:00:00", -1, 0.1, 101])
    closed.append(["Outside (OLD)", "12/01/2024 12:00:00", 20, 0, 303])
    if include_gold:
        closed.append(["Gold (Non Expiry) (GOLD)", "09/01/2024 15:00:00", 10, 0, 202])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class AccountStatementTest(unittest.TestCase):
    def test_analyzes_portfolio_and_modified_dietz_with_dated_flow(self):
        analysis = analyze_account_statement(analysis_workbook())

        self.assertEqual(analysis.total_profit_loss, 15)
        self.assertAlmostEqual(analysis.simple_roi, 0.10)
        self.assertAlmostEqual(analysis.modified_dietz_return, 0.12)
        self.assertGreater(analysis.annualized_roi, analysis.simple_roi)
        self.assertEqual(analysis.net_external_flows, 50)
        self.assertEqual(analysis.open_positions, 2)
        self.assertEqual(analysis.long_exposure, 80)
        self.assertEqual(analysis.short_exposure, 20)
        self.assertEqual(
            [(group.name, group.value) for group in analysis.exposure_by_type],
            [("Stocks", 80), ("CFD", 20)],
        )
        bridge = (
            analysis.beginning_unrealized_equity
            + analysis.net_external_flows
            + analysis.closed_positions_profit_loss
            + analysis.dividends
            + analysis.fees
            + analysis.other_performance
            + analysis.unrealized_profit_loss_change
        )
        self.assertAlmostEqual(bridge, analysis.ending_unrealized_equity)
        self.assertFalse(analysis.warnings)

    def test_undated_summary_flow_uses_midpoint_and_warns(self):
        analysis = analyze_account_statement(analysis_workbook(dated_deposit=False))

        self.assertEqual(len(analysis.cash_flows), 1)
        self.assertTrue(analysis.cash_flows[0].estimated_date)
        self.assertAlmostEqual(analysis.modified_dietz_return, 0.12)
        self.assertTrue(analysis.warnings)

    def test_date_range_reports_only_realized_activity(self):
        analysis = analyze_statement_range(
            analysis_workbook(),
            datetime(2024, 1, 7).date(),
            datetime(2024, 1, 9).date(),
        )

        self.assertEqual(analysis.realized_profit_loss, 6)
        self.assertEqual(analysis.closed_positions_profit_loss, 5)
        self.assertEqual(analysis.dividends, 2)
        self.assertEqual(analysis.fees, -1)
        self.assertEqual(analysis.net_external_flows, 0)
        self.assertEqual(analysis.estimated_beginning_equity, 150)
        self.assertEqual(analysis.estimated_ending_equity, 156)
        self.assertEqual(analysis.estimated_total_profit_loss, 6)
        self.assertAlmostEqual(analysis.estimated_roi, 0.04)
        self.assertAlmostEqual(analysis.estimated_modified_dietz_return, 0.04)
        self.assertEqual(analysis.holdings_snapshot_count, 0)
        self.assertTrue(analysis.valuation_warnings)
        self.assertEqual(
            [(point.day.day, point.cumulative_profit_loss) for point in analysis.daily_performance],
            [(7, 5), (8, 7), (9, 6)],
        )

    def test_date_range_estimates_unrealized_performance_from_holdings_snapshot(self):
        analysis = analyze_statement_range(
            analysis_workbook(holdings_snapshot=True),
            datetime(2024, 1, 7).date(),
            datetime(2024, 1, 9).date(),
        )

        self.assertAlmostEqual(analysis.estimated_beginning_equity, 175.6)
        self.assertAlmostEqual(analysis.estimated_ending_equity, 168.8)
        self.assertAlmostEqual(analysis.estimated_total_profit_loss, -6.8)
        self.assertAlmostEqual(analysis.estimated_roi, -6.8 / 175.6)
        self.assertEqual(analysis.holdings_snapshot_count, 1)
        self.assertEqual(analysis.max_boundary_anchor_distance_days, 2)
        self.assertEqual(
            [round(point.estimated_cumulative_profit_loss or 0, 2) for point in analysis.daily_performance],
            [5.0, 0.6, -6.8],
        )

    def test_date_range_validates_order_and_statement_boundaries(self):
        payload = analysis_workbook()
        with self.assertRaisesRegex(AccountStatementError, "must not be before"):
            analyze_statement_range(
                payload,
                datetime(2024, 1, 5).date(),
                datetime(2024, 1, 4).date(),
            )
        with self.assertRaisesRegex(AccountStatementError, "within the account statement"):
            analyze_statement_range(
                payload,
                datetime(2023, 12, 31).date(),
                datetime(2024, 1, 4).date(),
            )

    def test_return_helpers_reject_invalid_denominators_and_periods(self):
        start = datetime(2024, 1, 1)
        end = datetime(2024, 1, 11)
        flow = ExternalCashFlow(datetime(2024, 1, 6), 50, "Deposit")

        self.assertAlmostEqual(modified_dietz_return(100, 165, [flow], start, end), 0.12)
        self.assertIsNone(modified_dietz_return(-100, -50, [], start, end))
        self.assertIsNone(modified_dietz_return(100, 110, [], end, start))
        self.assertIsNone(annualize_return(-1, start, end))
        self.assertIsNone(annualize_return(0.1, end, start))

    def test_closed_position_contributions_are_grouped_by_asset_and_date(self):
        contributions = analyze_position_contributions(
            analysis_workbook(),
            datetime(2024, 1, 7).date(),
            datetime(2024, 1, 10).date(),
        )

        self.assertEqual(len(contributions), 1)
        self.assertEqual(contributions[0].asset, "Alpha (AAA)")
        self.assertEqual(contributions[0].closed_positions, 2)
        self.assertAlmostEqual(contributions[0].realized_profit_loss, 4)
        self.assertAlmostEqual(contributions[0].fees_and_dividends, -0.4)
        self.assertAlmostEqual(contributions[0].total_contribution, 4)

    def test_lists_tickers_and_excludes_linked_instrument_activity(self):
        payload = analysis_workbook(include_gold=True)

        self.assertIn("GOLD", list_statement_assets(payload))
        analysis = analyze_statement_range(
            payload,
            datetime(2024, 1, 7).date(),
            datetime(2024, 1, 9).date(),
            excluded_assets=("GOLD",),
        )
        contributions = analyze_position_contributions(
            payload,
            datetime(2024, 1, 7).date(),
            datetime(2024, 1, 10).date(),
            excluded_assets=("GOLD",),
        )

        self.assertEqual(analysis.realized_profit_loss, 6)
        self.assertEqual(analysis.closed_positions_profit_loss, 5)
        self.assertTrue(any("Excluded instruments" in warning for warning in analysis.valuation_warnings))
        self.assertEqual([item.asset for item in contributions], ["Alpha (AAA)"])

    def test_inspects_etoro_workbook_metadata_and_sheet_sizes(self):
        overview = inspect_account_statement(statement_workbook())

        self.assertEqual(overview.currency, "USD")
        self.assertEqual(overview.start_date, datetime(2023, 6, 1))
        self.assertEqual(overview.end_date, datetime(2023, 8, 31, 23, 59, 59))
        self.assertEqual(
            [(sheet.name, sheet.data_rows, sheet.columns) for sheet in overview.sheets],
            [
                ("Account Summary", 3, 2),
                ("Holdings", 2, 3),
                ("Glossary", 1, 2),
            ],
        )

    def test_reads_bounded_sheet_preview(self):
        preview = read_statement_sheet(statement_workbook(), "Holdings", max_rows=1)

        self.assertEqual(preview.columns, ("Asset", "Position ID", "Value in USD"))
        self.assertEqual(preview.rows, (("Apple", 123, 10.5),))
        self.assertEqual(preview.total_rows, 2)
        self.assertTrue(preview.truncated)

    def test_preview_keeps_columns_missing_from_the_first_row(self):
        preview = read_statement_sheet(statement_workbook(), "Glossary")

        self.assertEqual(preview.columns, ("Account Statement Glossary", "Column 2"))
        self.assertEqual(preview.rows, (("Term", "Meaning"),))

    def test_rejects_non_xlsx_and_workbook_without_account_summary(self):
        with self.assertRaisesRegex(AccountStatementError, "not an XLSX"):
            inspect_account_statement(b"not a workbook")
        with self.assertRaisesRegex(AccountStatementError, "missing sheet Account Summary"):
            inspect_account_statement(statement_workbook(include_summary=False))

    def test_rejects_unknown_sheet_and_invalid_preview_limit(self):
        payload = statement_workbook()
        with self.assertRaisesRegex(AccountStatementError, "not present"):
            read_statement_sheet(payload, "Closed Positions")
        with self.assertRaisesRegex(ValueError, "positive"):
            read_statement_sheet(payload, "Holdings", max_rows=0)


if __name__ == "__main__":
    unittest.main()
